"""Every writer produces a valid document from streamed rows; XLSX splits at the Excel limit."""

import csv
import io
import json
import xml.etree.ElementTree as ET

from openpyxl import load_workbook

from shared.enums import ExportFormat
from shared.exports.writers import (
    CsvWriter,
    GeoJsonWriter,
    GpxWriter,
    JsonWriter,
    XlsxWriter,
    make_writer,
)

COLUMNS = ["time", "latitude", "longitude", "altitude_m", "name", "attributes"]
ROWS = [
    {
        "time": "2026-04-01T02:00:00+02:00",
        "time_utc": "2026-04-01T00:00:00.000Z",
        "latitude": -24.9,
        "longitude": 31.5,
        "altitude_m": 300.5,
        "name": "Rhino 14",
        "attributes": {"fix_type": 3},
        "track_key": "a",
        "track_name": "Rhino 14",
    },
    {
        "time": "2026-04-01T02:10:00+02:00",
        "time_utc": "2026-04-01T00:10:00.000Z",
        "latitude": -24.91,
        "longitude": 31.51,
        "altitude_m": None,
        "name": 'Rhino "14"',
        "attributes": {},
        "track_key": "a",
        "track_name": "Rhino 14",
    },
    {
        "time": "2026-04-01T02:20:00+02:00",
        "time_utc": "2026-04-01T00:20:00.000Z",
        "latitude": -24.92,
        "longitude": 31.52,
        "altitude_m": 301.0,
        "name": "Elephant & co",
        "attributes": {"a": [1, 2]},
        "track_key": "b",
        "track_name": "Elephant & co",
    },
]
META = {"generator": "test"}


def _write(writer_cls, *args):
    stream = io.BytesIO()
    writer = writer_cls(stream, COLUMNS, *args)
    for row in ROWS:
        writer.write_row(row)
    writer.finish()
    return stream.getvalue()


def test_csv_quotes_and_serializes_nested_values():
    text = _write(CsvWriter).decode()
    rows = list(csv.reader(io.StringIO(text)))
    assert rows[0] == COLUMNS
    assert rows[2][4] == 'Rhino "14"' and rows[2][3] == ""
    assert json.loads(rows[1][5]) == {"fix_type": 3}


def test_json_is_one_document_with_metadata():
    document = json.loads(_write(JsonWriter, META))
    assert document["metadata"] == META and document["columns"] == COLUMNS
    assert len(document["rows"]) == 3 and document["rows"][2]["attributes"] == {"a": [1, 2]}
    assert "track_key" not in document["rows"][0]  # only listed columns are written


def test_geojson_points_with_properties():
    document = json.loads(_write(GeoJsonWriter, META))
    assert document["type"] == "FeatureCollection" and len(document["features"]) == 3
    feature = document["features"][0]
    assert feature["geometry"] == {"type": "Point", "coordinates": [31.5, -24.9]}
    assert feature["properties"]["name"] == "Rhino 14" and "latitude" not in feature["properties"]


def test_gpx_one_track_per_key_with_utc_times():
    root = ET.fromstring(_write(GpxWriter, META))
    ns = {"g": "http://www.topografix.com/GPX/1/1"}
    tracks = root.findall("g:trk", ns)
    assert [t.find("g:name", ns).text for t in tracks] == ["Rhino 14", "Elephant & co"]
    points = tracks[0].findall("g:trkseg/g:trkpt", ns)
    assert len(points) == 2 and points[0].get("lat") == "-24.9"
    assert points[0].find("g:time", ns).text == "2026-04-01T00:00:00.000Z"
    assert points[0].find("g:ele", ns).text == "300.5" and points[1].find("g:ele", ns) is None


def test_xlsx_splits_sheets_at_the_row_limit():
    stream = io.BytesIO()
    writer = XlsxWriter(stream, COLUMNS, max_rows=3)  # header plus two rows per sheet
    for row in ROWS:
        writer.write_row(row)
    writer.finish()
    workbook = load_workbook(io.BytesIO(stream.getvalue()), read_only=True)
    assert workbook.sheetnames == ["data", "data_2"] and writer.sheets == 2
    first = list(workbook["data"].iter_rows(values_only=True))
    second = list(workbook["data_2"].iter_rows(values_only=True))
    assert first[0] == tuple(COLUMNS) and second[0] == tuple(COLUMNS)
    assert len(first) == 3 and len(second) == 2
    assert first[1][4] == "Rhino 14" and json.loads(first[1][5]) == {"fix_type": 3}


def test_factory_covers_every_format():
    for export_format in ExportFormat:
        writer = make_writer(export_format, io.BytesIO(), COLUMNS, META)
        writer.write_row(ROWS[0])
        writer.finish()
